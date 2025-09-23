from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def freeze_header_and_first_column(ws: Worksheet) -> None:
    """Freeze the header row (row 1) and the first column (column A)."""
    ws.freeze_panes = "B2"


def freeze_header(ws: Worksheet) -> None:
    """Freeze only the header row (row 1)."""
    ws.freeze_panes = "A2"


def apply_auto_filter(ws: Worksheet, header_row: int = 1, max_col: int = 11) -> None:
    """Apply an auto filter from header to the last data row."""
    end_col_letter = get_column_letter(max_col)
    ws.auto_filter.ref = f"A{header_row}:{end_col_letter}{ws.max_row}"


def set_widths_from_cell(
    ws: Worksheet,
    max_col: int | None = None,
    padding: int = 2,
    min_width: int = 8,
) -> None:
    last_col = max_col or ws.max_column
    for col in range(1, last_col + 1):
        cell_max_width = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            length = len(str(cell.value)) if cell.value is not None else 10
            cell_max_width = max(cell_max_width, length + padding)
        ws.column_dimensions[get_column_letter(col)].width = cell_max_width
